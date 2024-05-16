import docx2txt
from docx import Document
from openpyxl import Workbook
import openpyxl

def extract_text_from_docx(docx_file):
    main_text = docx2txt.process(docx_file)
    doc = Document(docx_file)
    text_boxes_text = []
    for shape in doc.inline_shapes:
        if shape.type == 3:  # code text box
            text_boxes_text.append(shape.text.strip())
    return main_text, text_boxes_text
def create_excel_file(main_text, text_boxes, excel_file):
    wb = openpyxl.Workbook()
    ws = wb.active
    
    # Save main text
    main_text_lines = main_text.split('\n')
    for row, line in enumerate(main_text_lines, start=1):
        ws.cell(row=row, column=1, value=line)
    
    # Save text boxes
    current_row = len(main_text_lines) + 2  # Start after the main text
    for text_box in text_boxes:
        text_box_lines = text_box.split('\n')
        text_box_lines = [line.strip() for line in text_box_lines if line.strip()]
        for line in text_box_lines:
            ws.cell(row=current_row, column=1, value=line)
            current_row += 1
    
    wb.save(excel_file)
    print(f"Excel file created successfully at {excel_file}")
def remove_empty_rows(excel_file):
    wb = openpyxl.load_workbook(excel_file)
    ws = wb.active
    
    # Iterate through rows and remove empty ones
    rows_to_remove = []
    for row in ws.iter_rows():
        if all(cell.value is None or str(cell.value).strip() == "" for cell in row):
            rows_to_remove.append(row)
    
    for row in rows_to_remove:
        ws.delete_rows(row[0].row)
    
    # Save the modified Excel file
    wb.save(excel_file)

# Replace 'output.xlsx' with the name of your Excel file

# fichier word 
main_text, text_boxes_text = extract_text_from_docx(r'C:\Users\user\Desktop\python\G25057685_PLMACTION11471781_CHECKREPORT_DA_BACKCHECK_converted.docx')
excel_file = input("Enter the path for the Excel output file: ")

    # Iterate through .docx files in the folder
create_excel_file(main_text, text_boxes_text, excel_file)
remove_empty_rows(excel_file)
print("Main Text:")
print(main_text)

print("\nText Boxes Text:")
print(text_boxes_text)

